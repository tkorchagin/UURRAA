# coding=utf-8
from openpyxl import load_workbook

__author__ = 'tkorchagin'


def normalise(arr):
    # for item in arr:
    #     print item
    square_arr = [item ** 2 for item in arr]
    square_sum = sum(square_arr)
    norm_coefficient = square_sum ** 0.5

    if norm_coefficient == 0:
        return None

    norm_arr = [item / norm_coefficient for item in arr]
    return norm_arr


def self_vector(arr):
    mult = 1
    for el in arr:
        mult *= el

    power = 1.0 / len(arr)
    return mult ** power


def alternative_weight(arr):
    result = [float(item) / sum(arr) for item in arr]
    return result


def degree_of_preference(index, arr, bigger_better=True):
    result = []
    value = arr[index] + 0.05  # Защита от деления на 0
    for q in xrange(len(arr)):
        div = arr[q] + 0.05  # Защита от деления на 0
        if bigger_better:
            v = value / div
        else:
            v = div / value
        result.append(v)
    return result


def get_data(fn, sheet_name, column, from_row=0):
    wb = load_workbook(fn)

    sheet = wb.get_sheet_by_name(sheet_name)
    row_count = sheet.get_highest_row()

    column_name = sheet.cell(row=0, column=column).value

    data = []
    for q in xrange(1, row_count):
        if q < from_row:
            continue
        cell = sheet.cell(row=q, column=column)
        value = cell.value
        data.append(value)
    return data, column_name


def get_weights(data, bigger_better):
    norma = normalise(data)
    self_vectors = []
    for q in xrange(len(norma)):
        preference_arr = degree_of_preference(q, norma, bigger_better)
        s_vector = self_vector(preference_arr)
        self_vectors.append(s_vector)
    weights = alternative_weight(self_vectors)
    return weights


if __name__ == '__main__':
    fn = './tables/rcb_uurraa_data.xlsx'
    sheet_name = u'main'
    from_row = 3  # Исключил Москву

    column_indexes = range(1, 10)
    weight_arrs = []
    for column in column_indexes:
        data, column_name = get_data(fn, sheet_name, column, from_row)
        if column_name in [u'Расстояние по а/д', u'Арендная плата в регионе', ]:
                           # u'Процент сельского населения',
                           # u'Количество людей с доходом ниже среднего']:
            bigger_better = False
        else:
            bigger_better = True

        data_weights = get_weights(data, bigger_better)
        weight_arrs.append(data_weights)

    regions, _ = get_data(fn, u'main', 0, from_row)
    coefficients_values, coeff_arr = get_data(fn, u'coefficients', 1, 1)
    c_values = [coefficients_values[q - 1] for q in column_indexes]
    c_weights = get_weights(c_values, True)

    print '*'*80
    for w in c_weights:
        print w

    print '*' * 80
    for reg_index in xrange(len(regions)):
        region_name = regions[reg_index]
        region_utitity = 0
        for coeff_index in xrange(len(c_weights)):
            c_weight = c_weights[coeff_index]
            region_utitity += c_weight * weight_arrs[coeff_index][reg_index]
        print '"%s" -- %.5f' % (region_name, region_utitity)
        # print region_utitity
